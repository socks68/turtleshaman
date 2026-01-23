import json
from pathlib import Path
import openpyxl

XLSX_PATH = Path("data/twow_enh_model.xlsx")
SHEET = "Main(manual)"
OUT_PATH = Path("data/scenario_main_manual.json")

# Cell map based on what your sheet uses (labels in col A/C/E/G/I, values in col B/D/F/H/J etc)
INPUT_CELLS = {
    # Core combat inputs
    "bonus_strength": "B2",
    "melee_ap": "B3",
    "bonus_agility": "B4",
    "phys_crit": "B5",      # stored as decimal (0.04 = 4%)
    "phys_hit": "B6",       # stored as decimal (0.02 = 2%)

    "min_weapon_dmg": "F2",
    "max_weapon_dmg": "F3",
    "weapon_speed": "F4",

    "fight_duration_sec": "H2",
    "target_level": "H3",
    "attacking_behind": "H5",

    # Imbue / shield / fire totem selections
    "shield_type": "J2",    # e.g. "Lightning"
    "imbue": "J3",          # e.g. "Flametongue" / "Windfury"
    "fire_totem_mode": "J4" # e.g. "Nova+Searing"
}

# Outputs (cached formula results)
OUTPUT_CELLS = {
    "expected_dps": "G61",
    "expected_tps": "I61",
}

# Optional: pull the per-source DPS table area if you want breakdown parity.
# Your sheet uses rows in the 40s-50s for components; we'll export a block so we can inspect later.
BREAKDOWN_RANGE = ("I42", "L59")  # label/value/% style region in your sheet

def cell_to_val(ws, addr):
    return ws[addr].value

def export_block(ws, top_left, bottom_right):
    from openpyxl.utils import coordinate_to_tuple
    r1, c1 = coordinate_to_tuple(top_left)
    r2, c2 = coordinate_to_tuple(bottom_right)
    grid = []
    for r in range(r1, r2 + 1):
        row = []
        for c in range(c1, c2 + 1):
            row.append(ws.cell(r, c).value)
        grid.append(row)
    return grid

def main():
    wb = openpyxl.load_workbook(XLSX_PATH, data_only=True)
    ws = wb[SHEET]

    inputs = {k: cell_to_val(ws, addr) for k, addr in INPUT_CELLS.items()}
    outputs = {k: cell_to_val(ws, addr) for k, addr in OUTPUT_CELLS.items()}

    breakdown = export_block(ws, *BREAKDOWN_RANGE)

    payload = {
        "sheet": SHEET,
        "inputs": inputs,
        "outputs": outputs,
        "breakdown_block": {
            "range": f"{BREAKDOWN_RANGE[0]}:{BREAKDOWN_RANGE[1]}",
            "grid": breakdown
        }
    }

    OUT_PATH.write_text(json.dumps(payload, indent=2), encoding="utf-8")
    print(f"Wrote {OUT_PATH}")

if __name__ == "__main__":
    main()

